import openpyxl
import mysql.connector  

# Worksheet Details
workbook1 = openpyxl.load_workbook("C:/Users/Acer/Desktop/Employee_Data.xlsx")
worksheet1 = workbook1.active
rows = worksheet1.max_row
cols = worksheet1.max_column
print(rows,cols)
  
#Create the connection object   
myconn = mysql.connector.connect(host = "localhost", user = "root",passwd="1234", database = "employee_database")  
  
mycursor = myconn.cursor()
sql = "INSERT INTO employee_details (country) VALUES (%s)"

for i in range(2,rows+1):
    country = worksheet1.cell(row=i,column=3).value.strip()
    val =[country]
    mycursor.execute(sql, val)
    myconn.commit()
    # print(mycursor.rowcount, "record inserted.")

print("Query Executed....")