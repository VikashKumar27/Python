import openpyxl
import mysql.connector  

# Worksheet Details
workbook1 = openpyxl.load_workbook("C:/Users/Acer/Desktop/Employee_Data.xlsx")
worksheet1 = workbook1.active
rows = worksheet1.max_row
cols = worksheet1.max_column
print(rows,cols)
  
#Create the connection object   
myconn = mysql.connector.connect(host = "localhost", user = "root",passwd="1234", database = "employee_database")  
  
mycursor = myconn.cursor()

# mycursor.execute("CREATE TABLE Employee_Details (employee_id integer , name VARCHAR(255), address VARCHAR(255),country VARCHAR(255),company VARCHAR(255),Department varchar(255),salary VARCHAR(255))")
sql = "INSERT INTO employee_details (employee_id,name, address, country,company,Department, Salary) VALUES (%s, %s, %s,%s,%s,%s,%s)"
val = ("John", "Highway 21","India")
for i in range(2,rows+1):
    employee_id = worksheet1.cell(row=i,column=1).value
    name = worksheet1.cell(row=i,column=2).value.strip()
    address = worksheet1.cell(row=i,column=3).value.strip()
    country = worksheet1.cell(row=i,column=4).value.strip()
    company = worksheet1.cell(row=i,column=5).value.strip()
    Department = worksheet1.cell(row=i,column=6).value.strip()
    Salary = worksheet1.cell(row=i,column=7).value.strip()
    val =(employee_id,name,address,country,company,Department,Salary)
    mycursor.execute(sql, val)
    myconn.commit()
    # print(mycursor.rowcount, "record inserted.")

print("Successfully completed....")