import openpyxl
import mysql.connector  

# Worksheet Details
workbook1 = openpyxl.load_workbook("C:/Users/Acer/Desktop/Student_Details.xlsx")
worksheet1 = workbook1['Student_Personal_Data']
rows = worksheet1.max_row
cols = worksheet1.max_column
print(rows,cols)
  
#Create the connection object   
myconn = mysql.connector.connect(host = "localhost", user = "root",passwd="1234", database = "school")  
  
mycursor = myconn.cursor()

mycursor.execute("CREATE TABLE student_data (student_id int not null primary key, name VARCHAR(255), email VARCHAR(255),Age int)")
sql = "INSERT INTO student_data (student_id,name,email,age) VALUES (%s, %s, %s,%s)"
# val = ("John", "Highway 21","India")
for i in range(2,rows+1):
    student_id = worksheet1.cell(row=i,column=1).value
    name = worksheet1.cell(row=i,column=2).value.strip()
    email = worksheet1.cell(row=i,column=3).value.strip()
    age = worksheet1.cell(row=i,column=4).value
    val =(student_id,name,email,age)
    mycursor.execute(sql, val)
    myconn.commit()
    # print(mycursor.rowcount, "record inserted.")

print("Successfully completed....")