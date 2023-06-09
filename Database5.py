import openpyxl
import mysql.connector  

# Worksheet Details
workbook1 = openpyxl.load_workbook("C:/Users/Acer/Desktop/Student_Details.xlsx")
worksheet1 = workbook1['Student_Personal_Data']
rows = worksheet1.max_row
cols = worksheet1.max_column
print(rows,cols)
  
#Create the connection object   
myconn = mysql.connector.connect(host = "localhost", user = "root",passwd="1234", database = "school")  
  
mycursor = myconn.cursor()

# mycursor.execute("CREATE TABLE job_role (id int, Department VARCHAR(255), company VARCHAR(255),Job_Role VARCHAR(255),Experience VARCHAR(255),Salary varchar(255))")
sql = "INSERT INTO job_role (id,Department,company,Job_Role,Experience,salary) VALUES (%s, %s, %s,%s,%s,%s)"
# val = ("John", "Highway 21","India")
for i in range(2,rows+1):
    id = worksheet1.cell(row=i,column=1).value
    Department = worksheet1.cell(row=i,column=2).value.strip()
    Company = worksheet1.cell(row=i,column=3).value.strip()
    Job_Role = worksheet1.cell(row=i,column=4).value.strip()
    Experience = worksheet1.cell(row=i,column=5).value.strip()
    Salary = worksheet1.cell(row=i,column=6).value.strip()
    val =(id,Department,Company,Job_Role,Experience,Salary)
    mycursor.execute(sql, val)
    myconn.commit()
    # print(mycursor.rowcount, "record inserted.")

print("Successfully completed....")