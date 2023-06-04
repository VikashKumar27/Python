import openpyxl
import mysql.connector  

# Worksheet Details
workbook1 = openpyxl.load_workbook("C:/Users/Acer/Desktop/Employee_Data.xlsx")
worksheet1 = workbook1.active
rows = worksheet1.max_row
cols = worksheet1.max_column
print(rows,cols)
  
#Create the connection object   
myconn = mysql.connector.connect(host = "localhost", user = "root",passwd="1234", database = "employee_database")  
  
mycursor = myconn.cursor()

# Adding column in Table 

Query = "Alter Table Employee_Details ADD Continent VARCHAR(255)"
mycursor.execute(Query)


print("Query Executed....")